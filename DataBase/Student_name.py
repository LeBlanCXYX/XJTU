import requests
import openpyxl
from lxml import etree
# 创建一个新的excel工作簿
workbook=openpyxl.Workbook()
# 获取默认的工作表
sheet=workbook.active
sheet['B1']='SNAME'
url="https://ee.xjtu.edu.cn/info/1383/12513.htm"
response=requests.get(url)
# 将html页面内容转换为xml文档对象
html=etree.HTML(response.content)
datalist=html.xpath('//tr')
for item in datalist:
    name=item.xpath('./td[2]/p/text()')
    if len(name)>0:
        sheet.append([' ',f'{name[0]}'])
workbook.save('D:\code\python\icourese\Sname.xlsx')

import openpyxl
import random
from datetime import datetime, timedelta
import pandas as pd

def generate_student_id(existing_ids):
    exclude_ids = ['01032010', '01032023', '01032001', '01032005', 
                   '01032112', '03031011', '03031014', '03031051', 
                   '03031009', '03031033', '03031056']
    
    while True:
        student_id = ''.join(random.choices('0123456789', k=8))
        if student_id not in existing_ids and student_id not in exclude_ids:
            return student_id


def generate_gender():
    return random.choice(['男', '女'])

def generate_birthdate():
    start_date = datetime.strptime('2002-01-01', '%Y-%m-%d')
    end_date = datetime.strptime('2005-12-31', '%Y-%m-%d')
    random_date = start_date + timedelta(days=random.randint(0, (end_date - start_date).days))
    return random_date.strftime('%Y-%m-%d')

def generate_height(gender):
    if gender == '男':
        return round(random.uniform(1.6, 1.9), 2)
    elif gender == '女':
        return round(random.uniform(1.55, 1.8), 2)
    else:
        return None

def generate_dorm(assigned_dorms, gender):
    while True:
        building = random.choice(['东', '西'])
        
        if gender == '女':
            # 女生只能分配到楼层 1-5 和 11-15
            floor = random.choice([1, 2, 3, 4, 5, 11, 12, 13, 14, 15])
        else:
            # 男生可以分配到除了 1-5 和 11-15 外的其他楼层
            floor = random.randint(6, 10)  # 楼层范围从6到10
        
        fls = str(random.randint(1, 9))  # 确保是两位数，如09
        room = str(random.randint(1, 40)).zfill(2)  # 房间号，确保是两位数，如05

        room_number = f'{building} {floor} 舍 {fls}{room}'

        # 获取已分配的宿舍情况
        male_dorms = [dorm['dorm'] for dorm in assigned_dorms if dorm['gender'] == '男']
        female_dorms = [dorm['dorm'] for dorm in assigned_dorms if dorm['gender'] == '女']

        # 如果只有男性宿舍或只有女性宿舍，则后续只能分配相应性别
        if room_number in female_dorms and gender == '女':
            continue
        elif room_number in male_dorms and gender == '男':
            continue

        # 如果没有冲突的宿舍分配，则返回宿舍号和性别
        break

    return room_number, gender


def main():
    # 打开Excel文件
    wb = openpyxl.load_workbook(r'D:\code\python\icourese\origin_table\Sname.xlsx')
    ws = wb.active

    existing_ids = set()
    assigned_dorms = []

    # 生成学号、性别、出生日期、身高和宿舍号并写入Excel文件
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_col=1, values_only=True), start=1):
        student_name = row[0]
        student_id = generate_student_id(existing_ids)
        gender = generate_gender()
        birthdate = generate_birthdate()
        height = generate_height(gender)
        dorm, dorm_gender = generate_dorm(assigned_dorms, gender)
        existing_ids.add(student_id)
        assigned_dorms.append({'dorm': dorm, 'gender': dorm_gender})
        
        # 写入Excel文件，第一列为学号，第二列为性别，依次类推
        ws.cell(row=idx, column=1, value=student_id)
        ws.cell(row=idx, column=2, value=student_name)  # 调换了第一列和第二列的位置
        ws.cell(row=idx, column=3, value=gender)
        ws.cell(row=idx, column=4, value=birthdate)
        ws.cell(row=idx, column=5, value=height)
        ws.cell(row=idx, column=6, value=dorm)

    wb.save(r'D:\\code\\python\\icourese\\processed_table\\New_Processed_Sname_table.xlsx')
main()
