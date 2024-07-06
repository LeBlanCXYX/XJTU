from bs4 import BeautifulSoup
import requests
import os
from openpyxl import Workbook, load_workbook

# 文件路径
file_path = 'D:\\code\\python\\icourese\\Cname.xlsx'

# 创建或加载Excel文件
if os.path.exists(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    # 写入表头（如果需要）
    ws.append(['Course Name'])

# 请求头
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36',
    'cookie': 'imooc_uuid=86ab3b90-3f49-41aa-ab5e-52d53e209d85; imooc_isnew=1; imooc_isnew_ct=1717230404; tgw_l7_route=126bb7f6409ec5ff7744080c81a70225; sajssdk_2015_cross_new_user=1; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2218fd2e889f68f0-02cdb6893c548f8-26001c51-2359296-18fd2e889f75a8%22%2C%22first_id%22%3A%22%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMThmZDJlODg5ZjY4ZjAtMDJjZGI2ODkzYzU0OGY4LTI2MDAxYzUxLTIzNTkyOTYtMThmZDJlODg5Zjc1YTgifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%22%2C%22value%22%3A%22%22%7D%2C%22%24device_id%22%3A%2218fd2e889f68f0-02cdb6893c548f8-26001c51-2359296-18fd2e889f75a8%22%7D; Hm_lvt_f0cfcccd7b1393990c78efdeebff3968=1717230406; Hm_lpvt_f0cfcccd7b1393990c78efdeebff3968=1717230412; cvde=665adb44bd037-4'
}

# 循环访问第2页到第10页的内容
for page in range(1, 11):
    # 构造URL
    url = f'http://www.imooc.com/course/list/?page={page}'

    # 发送HTTP请求获取网页内容
    response = requests.get(url, headers=headers)

    # 创建BeautifulSoup对象
    soup = BeautifulSoup(response.text, 'html.parser')

    # 提取符合条件的<a>标签
    a_tags = soup.find_all('a', class_="item free")

    # 提取并写入data-title属性的值
    for tag in a_tags:
        data_title = tag.get('data-title')
        if data_title:
            print(data_title)
            ws.append([data_title])

# 保存Excel文件
wb.save(file_path)
print('数据爬取完成！')


import pandas as pd
import random
import openpyxl
from faker import Faker
from math import floor

# 创建一个新的Excel文件
file_path = r'D:\code\python\icourese\processed_table\Processed_Course.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Sheet1'

# 生成AA-BB
AA_values = ["CS", "EE", "ME", "CE", "BE", "AE", "EC", "FN", "MA", "MT", "PH", "CH", "BI", "AR", "DE", "PS", "ST", "GE", "HI", "LS"]
excluded_codes = ["CS-01", "CS-02", "CS-04", "CS-05", "EE-02", "EE-03", "CS-03", "EE-01"]
generated_codes = set(excluded_codes)

while len(generated_codes) < 100 + len(excluded_codes):
    AA = random.choice(AA_values)
    BB = f"{random.randint(1, 99):02}"
    code = f"{AA}-{BB}"
    if code not in generated_codes:
        generated_codes.add(code)

# 去掉排除的代码，只保留生成的100个代码
generated_codes -= set(excluded_codes)
generated_codes = list(generated_codes)[:100]

# 写入第一列
for i, code in enumerate(generated_codes, start=2):
    ws.cell(row=i, column=1, value=code)

# 读取现有Excel文件中的数据并写入新文件的第二列
origin_file_path = r'D:\code\python\icourese\origin_table\Cname.xlsx'
origin_data = pd.read_excel(origin_file_path)
second_column_data = origin_data.iloc[1:101, 0].tolist()  # 读取第2行到第101行的数据

for i, data in enumerate(second_column_data, start=2):
    ws.cell(row=i, column=2, value=data)

# 随机生成指定范围内的数字，并写入第三列和第四列
for i in range(2, 102):
    num = random.choice(range(24, 101, 4))
    decimal_part = num % 20 / 20.0  # 计算小数部分
    if decimal_part >= 0.5:
        adjusted_num = floor(num / 20.0) + 0.5
    else:
        adjusted_num = floor(num / 20.0)
    ws.cell(row=i, column=3, value=num)
    ws.cell(row=i, column=4, value=adjusted_num)

# 随机生成中文名字并写入第五列
faker = Faker('zh_CN')
for i in range(2, 102):
    name = faker.name()
    ws.cell(row=i, column=5, value=name)

# 保存Excel文件
wb.save(file_path)
print('数据处理完成！')
